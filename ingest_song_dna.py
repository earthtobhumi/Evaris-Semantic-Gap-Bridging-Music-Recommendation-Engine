from openpyxl import load_workbook
import pandas as pd
import sqlite3
import re
 
XLSX_PATH = "song_dna_finder.xlsx"
DB_PATH   = "song_dna.db"
TABLE     = "songs"
 
wb   = load_workbook(XLSX_PATH)
ws   = wb["song_data"]
rows = list(ws.iter_rows(values_only=True))
 
headers = [re.sub(r"[\s\(\)\-\/]+", "_", str(h).strip().lower()) for h in rows[0]]
data    = [r for r in rows[1:] if any(v is not None for v in r)]
df      = pd.DataFrame(data, columns=headers)
 
con = sqlite3.connect(DB_PATH)
df.to_sql(TABLE, con, if_exists="replace", index=False)
 
count = pd.read_sql(f"SELECT COUNT(*) as n FROM {TABLE}", con).iloc[0, 0]
print(f"✅ {count} rows written → {DB_PATH}")
print(df[["song", "artist"]].to_string())
con.close()
