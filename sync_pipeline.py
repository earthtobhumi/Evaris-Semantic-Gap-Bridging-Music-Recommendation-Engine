"""
sync_pipeline.py — EVARIS Master Sync Pipeline
================================================
Safely runs the full track-addition pipeline end-to-end:

  1. Validate Excel (song_dna_finder.xlsx)
  2. Ingest Excel -> SQLite (songs table)
  3. NLP embed -> SQLite (sentiment_embeddings table)
  4. Reddit blend -> enrich sentiment_embeddings in-place
  5. DSP batch -> SQLite (audio_features table), matched against audio/
  6. Rebuild ChromaDB from sentiment_embeddings
  7. Migrate songs + audio_features -> Supabase Postgres

Each stage is a checkpoint: if one fails, the script stops and tells you
exactly what to fix before re-running. Re-running is safe — every stage
is idempotent (replace/upsert, not append).

Usage:
    python sync_pipeline.py                  # full run
    python sync_pipeline.py --skip-reddit     # skip Reddit blending (faster, no crowd sentiment)
    python sync_pipeline.py --from-step 3     # resume from a specific step (1-7)
    python sync_pipeline.py --dry-run         # validate only, no writes
"""

import sys
import argparse
import sqlite3
import re
import time
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ============================================================
# CONFIG
# ============================================================
XLSX_PATH   = "song_dna_finder.xlsx"
DB_PATH     = "song_dna.db"
AUDIO_DIR   = "audio"
CHROMA_DIR  = "chroma_store"
TABLE_SONGS = "songs"

REQUIRED_HEADERS = {
    "SONG", "ARTIST", "LANGUAGE", "GENRE", "PRIMARY VIBE",
    "ENERGY SCORE", "SCENE DESCRIPTION"
}
SHEET_NAME = "song_data"
MIN_SCENE_LEN = 50

AUDIO_EXTS = {".mp3", ".wav", ".flac", ".m4a"}


def log(msg):
    print(msg, flush=True)


def fail(msg):
    log(f"\n❌ PIPELINE STOPPED: {msg}")
    log("Fix the issue above and re-run. No further steps were executed.")
    sys.exit(1)


# ============================================================
# STEP 1 — VALIDATE EXCEL
# ============================================================
def step1_validate_excel():
    log("\n" + "=" * 60)
    log("STEP 1 — Validating Excel file")
    log("=" * 60)

    if not Path(XLSX_PATH).exists():
        fail(f"'{XLSX_PATH}' not found in current directory.")

    wb = load_workbook(XLSX_PATH, read_only=True)
    if SHEET_NAME not in wb.sheetnames:
        fail(f"Sheet '{SHEET_NAME}' not found. Sheets present: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        fail("Sheet is empty.")

    headers = [str(h).strip() for h in rows[0]]
    missing_headers = REQUIRED_HEADERS - set(headers)
    if missing_headers:
        fail(f"Missing required headers: {missing_headers}")

    data = [r for r in rows[1:] if any(v is not None for v in r)]
    if not data:
        fail("No data rows found after header.")

    col_idx = {h: i for i, h in enumerate(headers)}
    issues = []

    # Missing required fields
    for i, row in enumerate(data, 2):
        for col in REQUIRED_HEADERS:
            val = row[col_idx[col]]
            if val is None or str(val).strip() == "":
                issues.append(f"  Row {i}: missing '{col}' (song: {row[col_idx['SONG']]})")

    # Duplicates
    song_names = [row[col_idx["SONG"]] for row in data if row[col_idx["SONG"]]]
    seen = {}
    for i, name in enumerate(song_names, 2):
        if name in seen:
            issues.append(f"  Row {i}: duplicate song '{name}' (first seen row {seen[name]})")
        else:
            seen[name] = i

    # Energy score range
    for i, row in enumerate(data, 2):
        val = row[col_idx["ENERGY SCORE"]]
        if val is not None and (not isinstance(val, (int, float)) or not (1 <= val <= 10)):
            issues.append(f"  Row {i}: energy score '{val}' out of range 1-10 (song: {row[col_idx['SONG']]})")

    # Scene description length
    for i, row in enumerate(data, 2):
        val = str(row[col_idx["SCENE DESCRIPTION"]] or "")
        if len(val.strip()) < MIN_SCENE_LEN:
            issues.append(f"  Row {i}: scene description too short ({len(val)} chars, song: {row[col_idx['SONG']]})")

    if issues:
        log(f"\n⚠️  {len(issues)} validation issue(s) found:")
        for issue in issues:
            log(issue)
        fail("Fix the Excel issues above before proceeding.")

    log(f"✅ {len(data)} rows validated — no issues found.")
    return len(data)


# ============================================================
# STEP 2 — INGEST EXCEL -> SQLITE
# ============================================================
def step2_ingest_excel():
    log("\n" + "=" * 60)
    log("STEP 2 — Ingesting Excel -> SQLite (songs table)")
    log("=" * 60)

    wb = load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))

    headers = [re.sub(r"[\s\(\)\-\/]+", "_", str(h).strip().lower()) for h in rows[0]]
    data = [r for r in rows[1:] if any(v is not None for v in r)]
    df = pd.DataFrame(data, columns=headers)

    con = sqlite3.connect(DB_PATH)
    df.to_sql(TABLE_SONGS, con, if_exists="replace", index=False)

    count = pd.read_sql(f"SELECT COUNT(*) as n FROM {TABLE_SONGS}", con).iloc[0, 0]
    con.close()

    log(f"✅ {count} rows written -> {DB_PATH} ({TABLE_SONGS} table)")

    if "energy_score" not in headers:
        log("⚠️  NOTE: SQLite column is 'energy_score' (from 'ENERGY SCORE' header).")
        log("    db.py's Song model uses 'energy_1_10' — these will NOT auto-map during")
        log("    Supabase migration since migrate_to_pg.py uses to_sql() directly")
        log("    (bypasses the ORM model). This is informational only — audio_features")
        log("    table (used by the live scorer) is unaffected.")

    return count


# ============================================================
# STEP 3 — NLP EMBEDDINGS
# ============================================================
def step3_nlp_embed():
    log("\n" + "=" * 60)
    log("STEP 3 — Generating NLP embeddings (MiniLM)")
    log("=" * 60)

    from sentence_transformers import SentenceTransformer

    MODEL = "paraphrase-multilingual-MiniLM-L12-v2"

    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT song, artist, scene_description, personal_notes FROM songs", con)

    df["combined_text"] = (
        df["scene_description"].fillna("") + " " + df["personal_notes"].fillna("")
    ).str.strip()

    log(f"🔤 Encoding {len(df)} tracks with {MODEL}...")
    model = SentenceTransformer(MODEL)
    embeddings = model.encode(df["combined_text"].tolist(), show_progress_bar=True)

    records = []
    for i, row in df.iterrows():
        records.append({
            "song": row["song"],
            "artist": row["artist"],
            "combined_text": row["combined_text"],
            "embedding_json": str(embeddings[i].tolist())
        })

    emb_df = pd.DataFrame(records)
    emb_df.to_sql("sentiment_embeddings", con, if_exists="replace", index=False)

    count = pd.read_sql("SELECT COUNT(*) as n FROM sentiment_embeddings", con).iloc[0, 0]
    con.close()

    log(f"✅ {count} embeddings written -> sentiment_embeddings table")
    log(f"📐 Embedding dimensions: {embeddings.shape[1]}")
    return count


# ============================================================
# STEP 4 — REDDIT CROWD SENTIMENT BLEND
# ============================================================
def step4_reddit_blend():
    log("\n" + "=" * 60)
    log("STEP 4 — Blending Reddit crowd sentiment")
    log("=" * 60)

    import requests
    import numpy as np
    from sentence_transformers import SentenceTransformer

    MODEL = "paraphrase-multilingual-MiniLM-L12-v2"
    MAX_COMMENTS = 30
    W_PERSONAL = 0.7
    W_CROWD = 0.3
    BASE_DELAY = 4          # seconds between requests, well under typical rate limits
    MAX_RETRIES = 4
    BACKOFF_BASE = 5         # 5s, 10s, 20s, 40s on consecutive 429s

    def fetch_reddit_comments(song, artist):
        """
        Returns (comments_text, status) where status is one of:
        'ok', 'empty' (no comments found, not rate-limited), 'rate_limited' (gave up after retries)
        """
        query = f"{song} {artist}"
        url = f"https://api.pullpush.io/reddit/search/comment/?q={requests.utils.quote(query)}&size={MAX_COMMENTS}"

        for attempt in range(MAX_RETRIES + 1):
            try:
                r = requests.get(url, timeout=30)
                if r.status_code == 429:
                    if attempt < MAX_RETRIES:
                        wait = BACKOFF_BASE * (2 ** attempt)
                        log(f"  ⏳ Rate limited (429) — retry {attempt+1}/{MAX_RETRIES} in {wait}s...")
                        time.sleep(wait)
                        continue
                    else:
                        return "", "rate_limited"
                r.raise_for_status()
                data = r.json().get("data", [])
                comments = [d["body"] for d in data if len(d.get("body", "")) > 20]
                text = " ".join(comments[:MAX_COMMENTS])
                return text, ("ok" if text.strip() else "empty")
            except Exception as e:
                if attempt < MAX_RETRIES:
                    wait = BACKOFF_BASE * (2 ** attempt)
                    log(f"  ⚠️  PullPush error for {song}: {e} — retry {attempt+1}/{MAX_RETRIES} in {wait}s...")
                    time.sleep(wait)
                    continue
                return "", "rate_limited"
        return "", "rate_limited"

    def blend(personal_vec, crowd_vec):
        blended = (W_PERSONAL * personal_vec) + (W_CROWD * crowd_vec)
        return blended / np.linalg.norm(blended)

    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT song, artist, embedding_json FROM sentiment_embeddings", con)
    model = SentenceTransformer(MODEL)

    updated = 0
    empty_results = []
    failed_after_retries = []

    for _, row in df.iterrows():
        log(f"🔍 Fetching Reddit comments: {row['song']} — {row['artist']}")
        crowd_text, status = fetch_reddit_comments(row["song"], row["artist"])

        if status == "rate_limited":
            failed_after_retries.append(row["song"])
            log(f"  ❌ Gave up after {MAX_RETRIES} retries — keeping personal embedding.")
            time.sleep(BASE_DELAY)
            continue

        if status == "empty":
            empty_results.append(row["song"])
            log(f"  ⏭️  No comments found (genuine), keeping personal embedding.")
            time.sleep(BASE_DELAY)
            continue

        personal_vec = np.array(eval(row["embedding_json"]))
        crowd_vec = model.encode([crowd_text])[0]
        blended_vec = blend(personal_vec, crowd_vec)

        con.execute(
            "UPDATE sentiment_embeddings SET embedding_json = ? WHERE song = ? AND artist = ?",
            (str(blended_vec.tolist()), row["song"], row["artist"])
        )
        con.commit()
        updated += 1
        log(f"  ✅ Blended embedding updated.")
        time.sleep(BASE_DELAY)

    con.close()
    log(f"\n🎵 {updated}/{len(df)} embeddings enriched with Reddit crowd sentiment.")
    if empty_results:
        log(f"⏭️  {len(empty_results)} song(s) genuinely had no Reddit comments.")
    if failed_after_retries:
        log(f"\n⚠️  SUMMARY: {len(failed_after_retries)} song(s) failed after {MAX_RETRIES} retries (still rate-limited):")
        for s in failed_after_retries:
            log(f"    - {s}")
        log("    Consider re-running --from-step 4 later, or increasing BASE_DELAY/BACKOFF_BASE.")
    return updated


# ============================================================
# STEP 5 — DSP BATCH EXTRACTION
# ============================================================
def step5_dsp_batch():
    log("\n" + "=" * 60)
    log("STEP 5 — Extracting DSP features from audio/")
    log("=" * 60)

    import librosa
    import numpy as np

    if not Path(AUDIO_DIR).exists():
        fail(f"'{AUDIO_DIR}/' directory not found.")

    pitch_classes = ['C', 'C#', 'D', 'D#', 'E', 'F', 'F#', 'G', 'G#', 'A', 'A#', 'B']
    major_profile = np.array([6.35, 2.23, 3.48, 2.33, 4.38, 4.09, 2.52, 5.19, 2.39, 3.66, 2.29, 2.88])
    minor_profile = np.array([6.35, 2.23, 3.48, 2.33, 4.38, 4.09, 2.52, 5.19, 2.39, 3.66, 2.29, 2.88])

    def extract(path):
        y, sr = librosa.load(path, mono=True)
        tempo = float(np.atleast_1d(librosa.beat.beat_track(y=y, sr=sr)[0])[0])
        chroma = librosa.feature.chroma_cqt(y=y, sr=sr).mean(axis=1)
        root = np.argmax(chroma)
        is_min = np.dot(chroma, np.roll(minor_profile, root)) < np.dot(chroma, np.roll(major_profile, root))
        key = f"{pitch_classes[root]} {'minor' if is_min else 'major'}"
        rms = float(librosa.feature.rms(y=y).mean())
        return tempo, key, rms

    def normalize(s):
        # Strip punctuation (parens, commas, apostrophes, etc.), collapse
        # whitespace/underscores to single spaces, lowercase. This lets
        # "Breathe (In the Air)" match "Breathe_In_The_Air" cleanly.
        s = s.strip().lower()
        s = re.sub(r"[_\-]+", " ", s)
        s = re.sub(r"[^\w\s]", "", s)  # drop (), ', ., etc.
        s = re.sub(r"\s+", " ", s).strip()
        return s

    con = sqlite3.connect(DB_PATH)
    songs_df = pd.read_sql("SELECT song, artist FROM songs", con)

    all_files = [p for ext in AUDIO_EXTS for p in Path(AUDIO_DIR).glob(f"*{ext}")]
    file_norms = [(p, normalize(p.stem)) for p in all_files]

    records = []
    unmatched = []
    ambiguous = []

    for _, row in songs_df.iterrows():
        db_norm = normalize(row["song"])

        # Pass 1: exact normalized match (preferred, unambiguous)
        exact = [p for p, fn in file_norms if fn == db_norm]

        if len(exact) == 1:
            matches = exact
        elif len(exact) > 1:
            # Multiple files normalize to the identical name — true ambiguity
            matches = exact
        else:
            # Pass 2: fuzzy substring fallback, only when no exact match exists
            matches = [
                p for p, fn in file_norms
                if db_norm in fn or fn in db_norm
            ]

        if not matches:
            unmatched.append(row["song"])
            log(f"⚠️  No file found for: {row['song']}")
            continue

        if len(matches) > 1:
            ambiguous.append((row["song"], [m.name for m in matches]))
            log(f"⚠️  Multiple matches for '{row['song']}': {[m.name for m in matches]} — using {matches[0].name}")

        try:
            tempo, key, rms = extract(matches[0])
            records.append({
                "song": row["song"], "artist": row["artist"],
                "tempo_bpm": round(tempo, 2), "key": key, "rms_energy": round(rms, 6)
            })
            log(f"✅ {row['song']} — {tempo:.1f} BPM | {key} | RMS {rms:.6f}")
        except Exception as e:
            log(f"❌ {row['song']} — {e}")

    if records:
        pd.DataFrame(records).to_sql("audio_features", con, if_exists="replace", index=False)
        log(f"\n🎵 {len(records)} tracks written -> audio_features table")

    con.close()

    if unmatched:
        log(f"\n⚠️  SUMMARY: {len(unmatched)} song(s) had NO audio file match:")
        for s in unmatched:
            log(f"    - {s}")
    if ambiguous:
        log(f"\n⚠️  SUMMARY: {len(ambiguous)} song(s) had AMBIGUOUS matches (check these manually):")
        for s, files in ambiguous:
            log(f"    - {s}: {files}")

    return len(records), unmatched, ambiguous


# ============================================================
# STEP 6 — REBUILD CHROMADB
# ============================================================
def step6_chroma_migrate():
    log("\n" + "=" * 60)
    log("STEP 6 — Rebuilding ChromaDB from sentiment_embeddings")
    log("=" * 60)

    import chromadb
    import shutil

    # Per risk register C1: delete chroma_store/ before rebuilding to avoid stale vectors
    if Path(CHROMA_DIR).exists():
        log(f"🗑️  Removing existing {CHROMA_DIR}/ to avoid stale/duplicate vectors...")
        shutil.rmtree(CHROMA_DIR)

    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT song, artist, combined_text, embedding_json FROM sentiment_embeddings", con)
    con.close()

    client = chromadb.PersistentClient(path=CHROMA_DIR)
    collection = client.get_or_create_collection(
        name="evaris_songs",
        metadata={"hnsw:space": "cosine"}
    )

    for _, row in df.iterrows():
        embedding = eval(row["embedding_json"])
        collection.upsert(
            ids=[f"{row['song']}_{row['artist']}"],
            embeddings=[embedding],
            documents=[row["combined_text"]],
            metadatas=[{"song": row["song"], "artist": row["artist"]}]
        )
        log(f"✅ Migrated: {row['song']} — {row['artist']}")

    final_count = collection.count()
    log(f"\n🎵 {final_count} tracks in ChromaDB -> {CHROMA_DIR}/")
    return final_count


# ============================================================
# STEP 7 — MIGRATE TO SUPABASE
# ============================================================
def step7_migrate_to_pg():
    log("\n" + "=" * 60)
    log("STEP 7 — Migrating songs + audio_features -> Supabase")
    log("=" * 60)

    from db import engine, init_db

    init_db()

    con = sqlite3.connect(DB_PATH)
    songs_df = pd.read_sql("SELECT * FROM songs", con)
    dsp_df = pd.read_sql("SELECT * FROM audio_features", con)
    con.close()

    songs_df.to_sql("songs", engine, if_exists="replace", index=False)
    log(f"✅ {len(songs_df)} songs migrated -> Supabase")

    dsp_df.to_sql("audio_features", engine, if_exists="replace", index=False)
    log(f"✅ {len(dsp_df)} audio features migrated -> Supabase")

    return len(songs_df), len(dsp_df)


# ============================================================
# MAIN
# ============================================================
STEPS = {
    1: ("Validate Excel", step1_validate_excel),
    2: ("Ingest Excel -> SQLite", step2_ingest_excel),
    3: ("NLP embeddings", step3_nlp_embed),
    4: ("Reddit blend", step4_reddit_blend),
    5: ("DSP batch extraction", step5_dsp_batch),
    6: ("Rebuild ChromaDB", step6_chroma_migrate),
    7: ("Migrate to Supabase", step7_migrate_to_pg),
}


def main():
    parser = argparse.ArgumentParser(description="EVARIS sync pipeline")
    parser.add_argument("--skip-reddit", action="store_true", help="Skip step 4 (Reddit blending)")
    parser.add_argument("--from-step", type=int, default=1, help="Resume from a specific step (1-7)")
    parser.add_argument("--dry-run", action="store_true", help="Only run step 1 (validation), no writes")
    args = parser.parse_args()

    log("🎵 EVARIS SYNC PIPELINE")
    log(f"Excel: {XLSX_PATH} | SQLite: {DB_PATH} | Audio: {AUDIO_DIR}/ | Chroma: {CHROMA_DIR}/")

    if args.dry_run:
        step1_validate_excel()
        log("\n✅ DRY RUN COMPLETE — no data was written.")
        return

    results = {}
    for step_num in range(args.from_step, 8):
        if step_num == 4 and args.skip_reddit:
            log("\n⏭️  SKIPPING Step 4 — Reddit blend (--skip-reddit flag set)")
            continue
        name, func = STEPS[step_num]
        results[step_num] = func()

    log("\n" + "=" * 60)
    log("✅ PIPELINE COMPLETE")
    log("=" * 60)
    log("Next: git add chroma_store/ requirements.txt && git commit && git push")
    log("Then reboot the Streamlit Cloud app to pick up the new Supabase data.")


if __name__ == "__main__":
    main()